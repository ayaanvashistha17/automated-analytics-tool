"""
Unit tests for ExcelAutomation module.
"""

import pytest
import pandas as pd
import os
import sys
import tempfile

# Add src to path
sys.path.append(os.path.join(os.path.dirname(__file__), "../src"))

from excel_automation import ExcelAutomation


@pytest.fixture
def automation():
    """Create ExcelAutomation instance."""
    return ExcelAutomation()


@pytest.fixture
def sample_data():
    """Create sample data for testing."""
    dates = pd.date_range(start="2024-01-01", periods=5, freq="D")
    data = pd.DataFrame(
        {
            "date": dates,
            "sales": [100, 110, 105, 120, 115],
            "revenue": [1000.0, 1100.0, 1050.0, 1200.0, 1150.0],
        }
    )
    return data


def test_excel_automation_initialization(automation):
    """Test ExcelAutomation initialization."""
    assert automation is not None
    assert hasattr(automation, "is_windows")
    assert isinstance(automation.is_windows, bool)


def test_create_vba_macro_file(automation, tmp_path):
    """Test VBA macro file creation."""
    # Create temporary file path
    output_path = os.path.join(tmp_path, "test_macros.txt")

    vba_code = automation.create_vba_macro_file(output_path=output_path)

    assert vba_code is not None
    assert isinstance(vba_code, str)
    assert len(vba_code) > 0

    # Check file was created
    assert os.path.exists(output_path)

    # Check file content
    with open(output_path, "r") as f:
        content = f.read()

    assert len(content) > 0
    assert "Automated Analytics Tool - VBA Macros" in content
    assert "Sub RefreshAllData" in content


def test_update_excel_data(automation, sample_data, tmp_path):
    """Test updating Excel data."""
    # Create temporary template file
    template_path = os.path.join(tmp_path, "test_template.xlsx")
    data_path = os.path.join(tmp_path, "test_data.csv")

    # Create a simple template
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Date"
    ws["B1"] = "Sales"
    ws["C1"] = "Revenue"
    wb.save(template_path)

    # Save data to CSV
    sample_data.to_csv(data_path, index=False)

    # Test updating Excel data
    automation.update_excel_data(template_path, data_path)

    # Verify the Excel file was updated
    from openpyxl import load_workbook

    wb_updated = load_workbook(template_path)
    ws_updated = wb_updated["Data"]

    # Check data was written (header + 5 rows)
    assert ws_updated.max_row == 6  # Header + 5 data rows

    wb_updated.close()


def test_vba_code_content(automation):
    """Test VBA code content structure."""
    vba_code = automation.create_vba_macro_file()

    # Check for essential macros
    assert "Sub RefreshAllData()" in vba_code
    assert "Sub GenerateDailyReport()" in vba_code

    # Check for error handling
    assert "On Error GoTo" in vba_code

    # Check for comments/documentation
    assert "'" in vba_code  # VBA comments


def test_cross_platform_compatibility(automation):
    """Test that automation works cross-platform."""
    # The module should work on both Windows and Mac
    assert hasattr(automation, "is_windows")

    # Should have the core methods
    assert hasattr(automation, "create_vba_macro_file")
    assert hasattr(automation, "update_excel_data")
    assert hasattr(automation, "automate_daily_process")


def test_automation_methods_exist(automation):
    """Test that all expected methods exist."""
    expected_methods = [
        "create_vba_macro_file",
        "update_excel_data",
        "automate_daily_process",
        "demonstrate_excel_automation",
    ]

    for method in expected_methods:
        assert hasattr(automation, method)
        assert callable(getattr(automation, method))


def test_vba_code_validity(automation):
    """Test that generated VBA code is syntactically valid."""
    vba_code = automation.create_vba_macro_file()

    # Check for proper VBA structure
    assert "Option Explicit" in vba_code
    assert "End Sub" in vba_code

    # Count Sub declarations vs End Sub
    sub_count = vba_code.count("Sub ")
    end_sub_count = vba_code.count("End Sub")
    assert sub_count == end_sub_count  # Should match


def test_template_creation(automation, tmp_path):
    """Test Excel template creation."""
    template_path = os.path.join(tmp_path, "test_template.xlsx")

    # Test creating template
    automation._create_template(template_path)

    assert os.path.exists(template_path)

    # Check template structure
    from openpyxl import load_workbook

    wb = load_workbook(template_path)

    assert "Data" in wb.sheetnames
    assert "Dashboard" in wb.sheetnames

    wb.close()


if __name__ == "__main__":
    # Run tests directly
    pytest.main([__file__, "-v", "--tb=short"])
