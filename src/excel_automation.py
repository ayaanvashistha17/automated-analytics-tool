"""
Excel automation module for VBA macro integration and Power BI synchronization.
Cross-platform compatible version.
"""

import os
import pandas as pd
from datetime import datetime
import logging
import shutil

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelAutomation:
    """Handle Excel automation with VBA macro integration (cross-platform)."""
    
    def __init__(self):
        """Initialize ExcelAutomation."""
        self.is_windows = os.name == 'nt'
        logger.info(f"ExcelAutomation initialized (Platform: {'Windows' if self.is_windows else 'Mac/Linux'})")
    
    def create_vba_macro_file(self, output_path="excel_files/macro_scripts/refresh_macros.txt"):
        """
        Create VBA macro code for documentation.
        
        Args:
            output_path (str): Path to save macro code
            
        Returns:
            str: VBA macro code
        """
        # Note: Using triple quotes with proper escaping for VBA code
        vba_code = '''\' ==============================================
\' Automated Analytics Tool - VBA Macros
\' ==============================================

Option Explicit

\' Main refresh macro
Sub RefreshAllData()
    \' Refresh all data connections and calculations
    On Error GoTo ErrorHandler
    
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    Application.EnableEvents = False
    
    \' Update timestamp
    ThisWorkbook.Worksheets("Dashboard").Range("LastUpdate").Value = Now
    
    \' Refresh all queries
    Dim qry As WorkbookQuery
    For Each qry In ThisWorkbook.Queries
        qry.Refresh
    Next qry
    
    \' Refresh all pivot tables
    Dim ws As Worksheet
    Dim pt As PivotTable
    For Each ws In ThisWorkbook.Worksheets
        For Each pt In ws.PivotTables
            pt.RefreshTable
        Next pt
    Next ws
    
    \' Calculate workbook
    ThisWorkbook.RefreshAll
    Application.Calculate
    
    \' Update Power BI connection if needed
    Call UpdatePowerBIConnection
    
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    Application.EnableEvents = True
    
    MsgBox "Data refresh completed successfully!", vbInformation
    Exit Sub
    
ErrorHandler:
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    Application.EnableEvents = True
    MsgBox "Error during refresh: " & Err.Description, vbCritical
End Sub

\' Update Power BI data source
Sub UpdatePowerBIConnection()
    On Error GoTo PBI_Error
    
    Dim dataSourcePath As String
    dataSourcePath = "C:\\Data\\Analytics\\daily_metrics.csv"
    
    \' Update Power Query connection
    Dim conn As WorkbookConnection
    For Each conn In ThisWorkbook.Connections
        If conn.Name Like "*PowerBI*" Then
            conn.OLEDBConnection.Connection = _
                "OLEDB;Provider=Microsoft.Mashup.OleDb.1;" & _
                "Data Source=$Workbook$;" & _
                "Location=" & dataSourcePath & ";" & _
                "Extended Properties="
        End If
    Next conn
    
    Exit Sub
    
PBI_Error:
    MsgBox "Power BI connection update failed: " & Err.Description, vbExclamation
End Sub

\' Generate daily report
Sub GenerateDailyReport()
    On Error GoTo ReportError
    
    Dim reportDate As String
    reportDate = Format(Date, "YYYY-MM-DD")
    
    \' Create report sheet
    Dim reportSheet As Worksheet
    Set reportSheet = ThisWorkbook.Worksheets.Add
    reportSheet.Name = "Report_" & reportDate
    
    \' Copy dashboard data to report
    ThisWorkbook.Worksheets("Dashboard").Range("A1:G20").Copy
    reportSheet.Range("A1").PasteSpecial Paste:=xlPasteAll
    
    \' Apply formatting
    With reportSheet
        .Range("A1").Value = "Daily Report - " & reportDate
        .Range("A1").Font.Bold = True
        .Range("A1").Font.Size = 14
        .Columns.AutoFit
    End With
    
    \' Save as separate file
    Dim savePath As String
    savePath = "C:\\Reports\\Daily_" & reportDate & ".xlsx"
    
    reportSheet.Copy
    ActiveWorkbook.SaveAs savePath, FileFormat:=xlOpenXMLWorkbook
    ActiveWorkbook.Close SaveChanges:=False
    
    MsgBox "Daily report generated: " & savePath, vbInformation
    Exit Sub
    
ReportError:
    MsgBox "Report generation failed: " & Err.Description, vbCritical
End Sub

\' Send email notification
Sub SendEmailNotification()
    On Error GoTo EmailError
    
    Dim OutlookApp As Object
    Dim MailItem As Object
    
    Set OutlookApp = CreateObject("Outlook.Application")
    Set MailItem = OutlookApp.CreateItem(0)
    
    With MailItem
        .To = "analytics-team@company.com"
        .CC = "manager@company.com"
        .Subject = "Daily Analytics Report - " & Format(Date, "YYYY-MM-DD")
        .Body = "The daily analytics report has been generated and is ready for review." & vbCrLf & vbCrLf & _
                "Key metrics have been updated and forecasts are available." & vbCrLf & vbCrLf & _
                "Please check the shared drive for the latest report."
        .Attachments.Add "C:\\Reports\\Daily_" & Format(Date, "YYYY-MM-DD") & ".xlsx"
        .Send
    End With
    
    Set MailItem = Nothing
    Set OutlookApp = Nothing
    
    MsgBox "Email notification sent!", vbInformation
    Exit Sub
    
EmailError:
    MsgBox "Email notification failed: " & Err.Description, vbExclamation
End Sub

\' Automated scheduled task
Sub ScheduledAutomation()
    \' This macro is called by Windows Task Scheduler
    Call RefreshAllData
    Call GenerateDailyReport
    Call SendEmailNotification
End Sub
'''
        
        # Save VBA code to file
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'w') as f:
            # Write the VBA code without the escape characters
            clean_vba = vba_code.replace("\\'", "'")
            f.write(clean_vba)
        
        logger.info(f"VBA macro code saved to: {output_path}")
        return clean_vba
    
    def update_excel_data(self, template_file, data_file):
        """
        Update Excel template with new data (cross-platform).
        
        Args:
            template_file (str): Path to Excel template
            data_file (str): Path to data file
        """
        try:
            # Load data
            data_df = pd.read_csv(data_file)
            
            # Update Excel file using openpyxl (cross-platform)
            from openpyxl import load_workbook
            
            wb = load_workbook(template_file)
            
            # Create or get Data sheet
            if "Data" in wb.sheetnames:
                ws = wb["Data"]
            else:
                ws = wb.create_sheet("Data")
            
            # Clear existing data (keep header)
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            
            # Write headers if needed
            if ws.max_row == 1 and ws['A1'].value is None:
                for col_idx, col_name in enumerate(data_df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
            
            # Write new data
            start_row = 2 if ws['A1'].value is not None else 1
            for row_idx, row in enumerate(data_df.itertuples(index=False), start_row):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Save workbook
            wb.save(template_file)
            logger.info(f"Updated {template_file} with data from {data_file}")
            
        except Exception as e:
            logger.error(f"Error updating Excel data: {str(e)}")
            raise
    
    def automate_daily_process(self, data_file, template_file="excel_files/report_template.xlsx", 
                              output_dir="reports/daily_reports"):
        """
        Run complete daily automation process (cross-platform).
        
        Args:
            data_file (str): Path to daily data file
            template_file (str): Path to Excel template
            output_dir (str): Output directory for reports
            
        Returns:
            str: Path to generated report
        """
        try:
            logger.info("Starting daily automation process...")
            
            # Step 1: Create template if it doesn't exist
            if not os.path.exists(template_file):
                self._create_template(template_file)
            
            # Step 2: Update data in template
            self.update_excel_data(template_file, data_file)
            
            # Step 3: Generate output file name
            report_date = datetime.now().strftime('%Y-%m-%d')
            output_file = f"{output_dir}/Daily_Report_{report_date}.xlsx"
            
            # Step 4: Copy template to output
            os.makedirs(output_dir, exist_ok=True)
            if os.path.exists(template_file):
                shutil.copy2(template_file, output_file)
                logger.info(f"Created report: {output_file}")
            
            # Step 5: Create VBA macro documentation
            self.create_vba_macro_file()
            
            logger.info(f"Daily automation completed. Report: {output_file}")
            return output_file
            
        except Exception as e:
            logger.error(f"Error in daily automation: {str(e)}")
            raise
    
    def _create_template(self, template_file):
        """Create Excel template if it doesn't exist."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create a new workbook
            wb = Workbook()
            
            # Remove default sheet
            default_sheet = wb.active
            wb.remove(default_sheet)
            
            # Create Data sheet
            data_ws = wb.create_sheet(title="Data")
            data_ws['A1'] = "Date"
            data_ws['B1'] = "Sales"
            data_ws['C1'] = "Revenue"
            data_ws['D1'] = "Users"
            data_ws['E1'] = "Conversion Rate"
            
            # Style the headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in data_ws["A1:E1"]:
                for c in cell:
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center")
            
            # Create Dashboard sheet
            dashboard_ws = wb.create_sheet(title="Dashboard")
            dashboard_ws['A1'] = "DAILY ANALYTICS DASHBOARD"
            dashboard_ws['A1'].font = Font(size=16, bold=True, color="1F497D")
            dashboard_ws.merge_cells('A1:E1')
            
            dashboard_ws['A3'] = "Last Updated:"
            dashboard_ws['B3'] = "=TODAY()"
            dashboard_ws['B3'].number_format = 'YYYY-MM-DD'
            
            # Create placeholder for metrics
            metrics = [
                ("Total Sales", "=SUM(Data!B:B)"),
                ("Total Revenue", "=SUM(Data!C:C)"),
                ("Average Users", "=AVERAGE(Data!D:D)"),
                ("Avg Conversion", "=AVERAGE(Data!E:E)")
            ]
            
            row = 5
            for label, formula in metrics:
                dashboard_ws.cell(row=row, column=1, value=label)
                dashboard_ws.cell(row=row, column=2, value=formula)
                row += 1
            
            # Save the template
            os.makedirs(os.path.dirname(template_file), exist_ok=True)
            wb.save(template_file)
            logger.info(f"Created template: {template_file}")
            
        except Exception as e:
            logger.error(f"Error creating template: {str(e)}")
            raise
    
    def demonstrate_excel_automation(self):
        """
        Demonstrate Excel automation capabilities.
        This shows what the tool can do and generates documentation.
        """
        print("=" * 60)
        print("Excel Automation Demonstration")
        print("=" * 60)
        print("\nThis module provides:")
        print("1. ✅ VBA Macro Code Generation")
        print("   - Creates ready-to-use VBA macros for Windows Excel")
        print("   - Includes: Data refresh, Power BI updates, Report generation")
        print("   - Saved to: excel_files/macro_scripts/refresh_macros.txt")
        
        print("\n2. ✅ Cross-Platform Excel Operations")
        print("   - Updates Excel files with latest data using openpyxl")
        print("   - Works on Mac, Linux, and Windows")
        print("   - Creates professional Excel templates")
        
        print("\n3. ✅ Power BI Integration Documentation")
        print("   - Shows how to connect Excel to Power BI")
        print("   - Documents the data refresh process")
        
        print("\n4. ✅ Daily Report Automation")
        print("   - Automatically updates Excel reports with new data")
        print("   - Creates timestamped reports")
        print("   - Maintains data history")
        
        print("\n" + "=" * 60)
        print("Implementation Notes:")
        print("- Actual VBA execution requires Windows with Excel installed")
        print("- The generated VBA code can be copied into Excel on Windows")
        print("- Excel data operations work cross-platform")
        print("- Power BI integration is documented for enterprise deployment")
        print("=" * 60)


# Simplified functions for direct use
def generate_vba_macros():
    """Generate VBA macro code for Excel automation."""
    automation = ExcelAutomation()
    return automation.create_vba_macro_file()


def update_excel_with_data(data_csv, excel_file):
    """Update Excel file with data from CSV."""
    automation = ExcelAutomation()
    automation.update_excel_data(excel_file, data_csv)
    print(f"✓ Updated {excel_file} with data from {data_csv}")


def create_daily_report(data_csv, output_dir="reports/daily_reports"):
    """Create daily report from data."""
    automation = ExcelAutomation()
    report_path = automation.automate_daily_process(data_csv)
    print(f"✓ Daily report created: {report_path}")
    return report_path


if __name__ == "__main__":
    # Example usage
    automation = ExcelAutomation()
    
    # Create VBA macro documentation
    vba_code = automation.create_vba_macro_file()
    print("✓ VBA macro code generated")
    
    # Demonstrate capabilities
    automation.demonstrate_excel_automation()
    
    print("\nUsage examples:")
    print("1. Generate VBA macros: generate_vba_macros()")
    print("2. Update Excel: update_excel_with_data('data.csv', 'template.xlsx')")
    print("3. Create report: create_daily_report('data/raw/daily_metrics.csv')")