import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Create template directory if it doesn't exist
os.makedirs("excel_files", exist_ok=True)

# Create a new workbook
wb = Workbook()

# Remove default sheet
default_sheet = wb.active
wb.remove(default_sheet)

# Create Data sheet
data_ws = wb.create_sheet(title="Data")

# Add headers
headers = ["Date", "Sales", "Revenue", "Users", "Conversion Rate"]
for col_idx, header in enumerate(headers, 1):
    data_ws.cell(row=1, column=col_idx, value=header)

# Style the headers
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for col in range(1, len(headers) + 1):
    cell = data_ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Create Dashboard sheet
dashboard_ws = wb.create_sheet(title="Dashboard")
dashboard_ws['A1'] = "DAILY ANALYTICS DASHBOARD"
dashboard_ws['A1'].font = Font(size=16, bold=True, color="1F497D")
dashboard_ws.merge_cells('A1:E1')

dashboard_ws['A3'] = "Last Updated:"
dashboard_ws['B3'] = "=TODAY()"
dashboard_ws['B3'].number_format = 'YYYY-MM-DD'

# Add sample metrics
metrics = [
    ("Total Sales", "=SUM(Data!B:B)"),
    ("Total Revenue", "=SUM(Data!C:C)"),
    ("Average Users", "=AVERAGE(Data!D:D)"),
    ("Avg Conversion Rate", "=AVERAGE(Data!E:E)"),
    ("Max Sales", "=MAX(Data!B:B)"),
    ("Min Sales", "=MIN(Data!B:B)")
]

row = 5
for label, formula in metrics:
    dashboard_ws.cell(row=row, column=1, value=label)
    dashboard_ws.cell(row=row, column=2, value=formula)
    row += 1

# Add some sample formatting
dashboard_ws.column_dimensions['A'].width = 20
dashboard_ws.column_dimensions['B'].width = 15

# Save the template
template_path = "excel_files/report_template.xlsx"
wb.save(template_path)
print(f"✓ Excel template created: {template_path}")
print("This template will be used by the automation tool.")
